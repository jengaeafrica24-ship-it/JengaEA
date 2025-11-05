from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from datetime import timedelta
import uuid
import os
from io import BytesIO

from .models import Report, ReportTemplate, ReportShare
from .serializers import (
    ReportSerializer, ReportCreateSerializer, ReportTemplateSerializer,
    ReportShareSerializer, ReportGenerationSerializer
)
from estimates.models import Estimate
from subscriptions.views import record_usage


class ReportListView(generics.ListCreateAPIView):
    """List and create reports"""
    
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return Report.objects.all()
        return Report.objects.filter(user=user)


class ReportDetailView(generics.RetrieveDestroyAPIView):
    """Retrieve or delete a report"""
    
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return Report.objects.all()
        return Report.objects.filter(user=user)


class ReportTemplateListView(generics.ListAPIView):
    """List available report templates"""
    
    queryset = ReportTemplate.objects.filter(is_active=True)
    serializer_class = ReportTemplateSerializer
    permission_classes = [IsAuthenticated]


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_report(request):
    """Generate a new report"""
    
    serializer = ReportGenerationSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Get estimate
        estimate = get_object_or_404(Estimate, id=serializer.validated_data['estimate_id'], user=request.user)
        
        # Check user's subscription quota
        usage_response = record_usage(request)
        if usage_response.status_code == 403:
            return usage_response
        
        # Generate report based on format
        report_type = serializer.validated_data['report_type']
        format_type = serializer.validated_data['format']
        title = serializer.validated_data['title']
        
        # Create report record
        report = Report.objects.create(
            user=request.user,
            estimate=estimate,
            report_type=report_type,
            format=format_type,
            title=title,
            description=serializer.validated_data.get('description', ''),
            file_path='',  # Will be set after file generation
            is_public=False
        )
        
        # Generate file based on format
        if format_type == 'pdf':
            file_content = generate_pdf_report(estimate, serializer.validated_data)
            file_extension = 'pdf'
        elif format_type == 'excel':
            file_content = generate_excel_report(estimate, serializer.validated_data)
            file_extension = 'xlsx'
        else:  # csv
            file_content = generate_csv_report(estimate, serializer.validated_data)
            file_extension = 'csv'
        
        # Save file
        filename = f"report_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}"
        file_path = os.path.join('reports', filename)
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'wb') as f:
            f.write(file_content)
        
        # Update report with file information
        report.file_path = file_path
        report.file_size = len(file_content)
        report.save()
        
        return Response({
            'message': 'Report generated successfully',
            'report': ReportSerializer(report).data,
            'download_url': f"/api/reports/{report.id}/download/"
        })
        
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def generate_pdf_report(estimate, options):
    """Generate PDF report"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Content
    story = []
    
    # Title
    story.append(Paragraph(options['title'], title_style))
    story.append(Spacer(1, 12))
    
    # Project Information
    project_info = [
        ['Project Name:', estimate.project_name],
        ['Project Type:', estimate.project_type.name],
        ['Location:', estimate.location.county_name],
        ['Total Area:', f"{estimate.total_area} sqm"],
        ['Generated Date:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    project_table = Table(project_info, colWidths=[2*inch, 4*inch])
    project_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
    ]))
    
    story.append(project_table)
    story.append(Spacer(1, 20))
    
    # Cost Summary
    story.append(Paragraph("Cost Summary", styles['Heading2']))
    story.append(Spacer(1, 12))
    
    cost_data = [
        ['Item', 'Amount (USD)'],
        ['Base Cost', f"${estimate.adjusted_cost_per_sqm * estimate.total_area:,.2f}"],
        ['Contingency ({estimate.contingency_percentage}%)', f"${estimate.contingency_amount:,.2f}"],
        ['', ''],
        ['Total Estimated Cost', f"${estimate.total_estimated_cost + estimate.contingency_amount:,.2f}"]
    ]
    
    cost_table = Table(cost_data, colWidths=[3*inch, 2*inch])
    cost_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    story.append(cost_table)
    story.append(Spacer(1, 20))
    
    # Estimate Items
    if estimate.items.exists():
        story.append(Paragraph("Detailed Breakdown", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        items_data = [['Category', 'Item', 'Quantity', 'Unit', 'Unit Price', 'Total Price']]
        for item in estimate.items.all():
            items_data.append([
                item.get_category_display(),
                item.name,
                str(item.quantity),
                item.unit,
                f"${item.unit_price:,.2f}",
                f"${item.total_price:,.2f}"
            ])
        
        items_table = Table(items_data, colWidths=[1*inch, 2*inch, 0.8*inch, 0.8*inch, 1*inch, 1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(items_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(estimate, options):
    """Generate Excel report"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Estimate"
    
    # Title
    ws['A1'] = options['title']
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    # Project Information
    row = 3
    project_info = [
        ['Project Name:', estimate.project_name],
        ['Project Type:', estimate.project_type.name],
        ['Location:', estimate.location.county_name],
        ['Total Area:', f"{estimate.total_area} sqm"],
        ['Generated Date:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for info in project_info:
        ws[f'A{row}'] = info[0]
        ws[f'B{row}'] = info[1]
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    row += 1
    
    # Cost Summary
    ws[f'A{row}'] = "Cost Summary"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    row += 2
    
    cost_data = [
        ['Item', 'Amount (USD)'],
        ['Base Cost', estimate.adjusted_cost_per_sqm * estimate.total_area],
        [f'Contingency ({estimate.contingency_percentage}%)', estimate.contingency_amount],
        ['', ''],
        ['Total Estimated Cost', estimate.total_estimated_cost + estimate.contingency_amount]
    ]
    
    for i, cost in enumerate(cost_data):
        ws[f'A{row}'] = cost[0]
        ws[f'B{row}'] = cost[1] if isinstance(cost[1], (int, float)) else cost[1]
        if i == 0 or i == len(cost_data) - 1:  # Header and total rows
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
        row += 1
    
    # Detailed breakdown
    if estimate.items.exists():
        row += 2
        ws[f'A{row}'] = "Detailed Breakdown"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Headers
        headers = ['Category', 'Item', 'Quantity', 'Unit', 'Unit Price', 'Total Price']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        row += 1
        
        # Items
        for item in estimate.items.all():
            ws[f'A{row}'] = item.get_category_display()
            ws[f'B{row}'] = item.name
            ws[f'C{row}'] = item.quantity
            ws[f'D{row}'] = item.unit
            ws[f'E{row}'] = item.unit_price
            ws[f'F{row}'] = item.total_price
            row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_csv_report(estimate, options):
    """Generate CSV report"""
    import csv
    
    buffer = BytesIO()
    
    # Write CSV content
    writer = csv.writer(buffer)
    
    # Title
    writer.writerow([options['title']])
    writer.writerow([])
    
    # Project Information
    writer.writerow(['Project Information'])
    writer.writerow(['Project Name', estimate.project_name])
    writer.writerow(['Project Type', estimate.project_type.name])
    writer.writerow(['Location', estimate.location.county_name])
    writer.writerow(['Total Area', f"{estimate.total_area} sqm"])
    writer.writerow(['Generated Date', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])
    
    # Cost Summary
    writer.writerow(['Cost Summary'])
    writer.writerow(['Item', 'Amount (USD)'])
    writer.writerow(['Base Cost', estimate.adjusted_cost_per_sqm * estimate.total_area])
    writer.writerow([f'Contingency ({estimate.contingency_percentage}%)', estimate.contingency_amount])
    writer.writerow([])
    writer.writerow(['Total Estimated Cost', estimate.total_estimated_cost + estimate.contingency_amount])
    writer.writerow([])
    
    # Detailed breakdown
    if estimate.items.exists():
        writer.writerow(['Detailed Breakdown'])
        writer.writerow(['Category', 'Item', 'Quantity', 'Unit', 'Unit Price', 'Total Price'])
        
        for item in estimate.items.all():
            writer.writerow([
                item.get_category_display(),
                item.name,
                item.quantity,
                item.unit,
                item.unit_price,
                item.total_price
            ])
    
    buffer.seek(0)
    return buffer.getvalue()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_report(request, report_id):
    """Download a report file"""
    
    try:
        report = get_object_or_404(Report, id=report_id, user=request.user)
        
        if not os.path.exists(report.file_path):
            return Response(
                {'error': 'Report file not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Increment download count
        report.download_count += 1
        report.save()
        
        # Return file
        response = FileResponse(
            open(report.file_path, 'rb'),
            as_attachment=True,
            filename=f"{report.title}.{report.format}"
        )
        
        return response
        
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def share_report(request, report_id):
    """Share a report via email"""
    
    try:
        report = get_object_or_404(Report, id=report_id, user=request.user)
        
        shared_with_email = request.data.get('email')
        shared_with_name = request.data.get('name', '')
        expires_days = request.data.get('expires_days', 30)
        
        if not shared_with_email:
            return Response(
                {'error': 'Email is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate access token
        access_token = str(uuid.uuid4())
        expires_at = timezone.now() + timedelta(days=expires_days)
        
        # Create share record
        share = ReportShare.objects.create(
            report=report,
            shared_with_email=shared_with_email,
            shared_with_name=shared_with_name,
            access_token=access_token,
            expires_at=expires_at,
            created_by=request.user
        )
        
        # TODO: Send email with share link
        
        return Response({
            'message': 'Report shared successfully',
            'share_token': access_token,
            'expires_at': expires_at,
            'share_url': f"/shared-report/{access_token}"
        })
        
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([])  # No authentication required for shared reports
def shared_report(request, access_token):
    """Access a shared report"""
    
    try:
        share = get_object_or_404(
            ReportShare, 
            access_token=access_token, 
            is_active=True
        )
        
        # Check if share has expired
        if timezone.now() > share.expires_at:
            return Response(
                {'error': 'This shared report has expired'}, 
                status=status.HTTP_410_GONE
            )
        
        return Response(
            ReportSerializer(share.report).data
        )
        
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )



